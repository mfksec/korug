"""Data export API endpoints."""
import json
import logging
import re
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Response, status
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill

from korug.db import get_db
from korug.auth_utils import get_current_user
from korug.audit import log_audit_event, AuditEvent
from korug.models import Domain, Subdomain, Vulnerability

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/xlsx/{domain_id}")
def export_to_xlsx(
    domain_id: int, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export domain scan results to XLSX file."""
    domain = db.query(Domain).filter(Domain.id == domain_id).first()
    if not domain:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Domain with id {domain_id} not found",
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Sheet 1: Subdomains and DNS Records
    ws_subdomains = wb.create_sheet("Subdomains")
    headers = ["Subdomain", "A Records", "AAAA Records", "CNAME", "MX Records", "NS Records", 
               "First Discovered", "Last Seen"]
    ws_subdomains.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_subdomains[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    subdomains = db.query(Subdomain).filter(Subdomain.domain_id == domain_id).all()
    for subdomain in subdomains:
        a_records = json.loads(subdomain.a_records) if subdomain.a_records else []
        aaaa_records = json.loads(subdomain.aaaa_records) if subdomain.aaaa_records else []
        mx_records = json.loads(subdomain.mx_records) if subdomain.mx_records else []
        ns_records = json.loads(subdomain.ns_records) if subdomain.ns_records else []
        
        row = [
            subdomain.subdomain,
            ", ".join(a_records),
            ", ".join(aaaa_records),
            subdomain.cname_record or "",
            ", ".join(mx_records),
            ", ".join(ns_records),
            subdomain.first_discovered.isoformat() if subdomain.first_discovered else "",
            subdomain.last_seen.isoformat() if subdomain.last_seen else "",
        ]
        ws_subdomains.append(row)
    
    # Adjust column widths
    ws_subdomains.column_dimensions["A"].width = 30
    ws_subdomains.column_dimensions["B"].width = 25
    ws_subdomains.column_dimensions["C"].width = 25
    ws_subdomains.column_dimensions["D"].width = 25
    ws_subdomains.column_dimensions["E"].width = 25
    ws_subdomains.column_dimensions["F"].width = 25
    ws_subdomains.column_dimensions["G"].width = 25
    ws_subdomains.column_dimensions["H"].width = 25
    
    # Sheet 2: Vulnerabilities
    ws_vulns = wb.create_sheet("Vulnerabilities")
    headers = ["Subdomain", "Vulnerability Type", "Confidence Score", "Details", "Found At", "False Positive"]
    ws_vulns.append(headers)
    
    # Style header row
    for cell in ws_vulns[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    vulnerabilities = db.query(Vulnerability).filter(Vulnerability.domain_id == domain_id).all()
    for vuln in vulnerabilities:
        subdomain = db.query(Subdomain).filter(Subdomain.id == vuln.subdomain_id).first()
        
        row = [
            subdomain.subdomain if subdomain else "",
            vuln.vuln_type,
            vuln.confidence_score,
            vuln.details or "",
            vuln.found_at.isoformat() if vuln.found_at else "",
            "Yes" if vuln.is_false_positive else "No",
        ]
        ws_vulns.append(row)
    
    # Adjust column widths
    ws_vulns.column_dimensions["A"].width = 30
    ws_vulns.column_dimensions["B"].width = 25
    ws_vulns.column_dimensions["C"].width = 15
    ws_vulns.column_dimensions["D"].width = 40
    ws_vulns.column_dimensions["E"].width = 25
    ws_vulns.column_dimensions["F"].width = 15
    
    # Serialize the workbook to bytes and return it as the response body.
    # (FileResponse expects a path on disk — passing it an in-memory buffer
    # raised a 500 for every export, regardless of whether the domain had data.)
    output = BytesIO()
    wb.save(output)
    data = output.getvalue()

    log_audit_event(
        AuditEvent.EXPORT_INITIATED,
        user=current_user['sub'],
        resource_type="domain",
        resource_id=domain_id,
        details={"domain_name": domain.domain_name, "format": "xlsx"}
    )

    # Sanitize the domain name for the download filename (it's user-supplied).
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", domain.domain_name) or "domain"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_report.xlsx"'},
    )
